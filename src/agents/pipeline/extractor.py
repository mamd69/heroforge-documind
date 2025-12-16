#!/usr/bin/env python3
"""
Document Content Extractor Agent

Extracts text content and metadata from various document formats:
- PDF (using PyPDF2 and pdfplumber as fallback)
- DOCX (using python-docx)
- XLSX (using openpyxl)
- TXT (plain text)
- MD (markdown)

Usage:
    python extractor.py <file_path>

Returns JSON with:
    - success: bool
    - file_path: str
    - title: str (from first heading or filename)
    - content: str
    - file_type: str
    - size: int (bytes)
    - error: str (if success=False)
"""

import json
import sys
from pathlib import Path
from typing import Dict, Any, Optional
import re


def extract_title_from_content(content: str, file_path: Path) -> str:
    """
    Extract title from content (first heading) or use filename as fallback.

    Args:
        content: Document content text
        file_path: Path to the file

    Returns:
        Extracted or inferred title
    """
    # Try to find first markdown heading
    md_heading = re.search(r'^#\s+(.+)$', content, re.MULTILINE)
    if md_heading:
        return md_heading.group(1).strip()

    # Try to find first line that looks like a title (short, non-empty)
    lines = content.strip().split('\n')
    for line in lines[:5]:  # Check first 5 lines
        line = line.strip()
        if line and len(line) < 100 and not line.startswith(('http', 'www', '//', '#')):
            return line

    # Fallback to filename without extension
    return file_path.stem


def extract_pdf(file_path: Path) -> Dict[str, Any]:
    """
    Extract content from PDF using PyPDF2, fallback to pdfplumber.

    Args:
        file_path: Path to PDF file

    Returns:
        Dictionary with extracted content and metadata
    """
    content = ""

    try:
        # Try PyPDF2 first (faster)
        import PyPDF2

        with open(file_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    content += text + "\n"

        # If PyPDF2 didn't extract much, try pdfplumber
        if len(content.strip()) < 100:
            import pdfplumber

            content = ""
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        content += text + "\n"

    except ImportError as e:
        return {
            "success": False,
            "error": f"Missing PDF library: {str(e)}. Install with: pip install PyPDF2 pdfplumber"
        }
    except Exception as e:
        return {
            "success": False,
            "error": f"PDF extraction failed: {str(e)}"
        }

    if not content.strip():
        return {
            "success": False,
            "error": "No text content extracted from PDF"
        }

    return {
        "success": True,
        "content": content.strip()
    }


def extract_docx(file_path: Path) -> Dict[str, Any]:
    """
    Extract content from DOCX using python-docx.

    Args:
        file_path: Path to DOCX file

    Returns:
        Dictionary with extracted content and metadata
    """
    try:
        from docx import Document

        doc = Document(file_path)
        content = "\n\n".join([paragraph.text for paragraph in doc.paragraphs if paragraph.text.strip()])

        if not content.strip():
            return {
                "success": False,
                "error": "No text content found in DOCX file"
            }

        return {
            "success": True,
            "content": content.strip()
        }

    except ImportError:
        return {
            "success": False,
            "error": "Missing python-docx library. Install with: pip install python-docx"
        }
    except Exception as e:
        return {
            "success": False,
            "error": f"DOCX extraction failed: {str(e)}"
        }


def extract_xlsx(file_path: Path) -> Dict[str, Any]:
    """
    Extract content from XLSX using openpyxl.

    Args:
        file_path: Path to XLSX file

    Returns:
        Dictionary with extracted content and metadata
    """
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(file_path, data_only=True)
        content_parts = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            content_parts.append(f"Sheet: {sheet_name}\n")

            # Extract all non-empty cells
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    content_parts.append(row_text)

            content_parts.append("")  # Blank line between sheets

        content = "\n".join(content_parts).strip()

        if not content:
            return {
                "success": False,
                "error": "No content found in XLSX file"
            }

        return {
            "success": True,
            "content": content
        }

    except ImportError:
        return {
            "success": False,
            "error": "Missing openpyxl library. Install with: pip install openpyxl"
        }
    except Exception as e:
        return {
            "success": False,
            "error": f"XLSX extraction failed: {str(e)}"
        }


def extract_text(file_path: Path) -> Dict[str, Any]:
    """
    Extract content from plain text or markdown files.

    Args:
        file_path: Path to text/markdown file

    Returns:
        Dictionary with extracted content and metadata
    """
    try:
        # Try UTF-8 first, then fallback to other encodings
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        content = None

        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as file:
                    content = file.read()
                break
            except UnicodeDecodeError:
                continue

        if content is None:
            return {
                "success": False,
                "error": "Failed to decode text file with supported encodings"
            }

        if not content.strip():
            return {
                "success": False,
                "error": "File is empty"
            }

        return {
            "success": True,
            "content": content.strip()
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"Text extraction failed: {str(e)}"
        }


def extract_document(file_path: str) -> Dict[str, Any]:
    """
    Main extraction function that routes to appropriate extractor based on file type.

    Args:
        file_path: Path to document file

    Returns:
        Dictionary with extraction results:
        - success: bool
        - file_path: str
        - title: str
        - content: str
        - file_type: str
        - size: int (bytes)
        - error: str (if success=False)
    """
    path = Path(file_path)

    # Validate file exists
    if not path.exists():
        return {
            "success": False,
            "file_path": str(file_path),
            "error": f"File not found: {file_path}"
        }

    if not path.is_file():
        return {
            "success": False,
            "file_path": str(file_path),
            "error": f"Not a file: {file_path}"
        }

    # Get file metadata
    file_type = path.suffix.lower().lstrip('.')
    file_size = path.stat().st_size

    # Route to appropriate extractor
    extractors = {
        'pdf': extract_pdf,
        'docx': extract_docx,
        'xlsx': extract_xlsx,
        'xls': extract_xlsx,
        'txt': extract_text,
        'md': extract_text,
    }

    extractor = extractors.get(file_type)
    if not extractor:
        return {
            "success": False,
            "file_path": str(file_path),
            "file_type": file_type,
            "size": file_size,
            "error": f"Unsupported file type: {file_type}. Supported: {', '.join(extractors.keys())}"
        }

    # Extract content
    result = extractor(path)

    # Add metadata to result
    result["file_path"] = str(path.resolve())
    result["file_type"] = file_type
    result["size"] = file_size

    if result["success"]:
        # Extract title from content
        result["title"] = extract_title_from_content(result["content"], path)

    return result


def main():
    """
    CLI entry point for standalone execution.

    Usage:
        python extractor.py <file_path>

    Outputs JSON to stdout.
    """
    if len(sys.argv) < 2:
        print(json.dumps({
            "success": False,
            "error": "Usage: python extractor.py <file_path>"
        }), file=sys.stderr)
        sys.exit(1)

    file_path = sys.argv[1]
    result = extract_document(file_path)

    # Output JSON to stdout
    print(json.dumps(result, indent=2))

    # Exit with appropriate code
    sys.exit(0 if result["success"] else 1)


if __name__ == "__main__":
    main()
